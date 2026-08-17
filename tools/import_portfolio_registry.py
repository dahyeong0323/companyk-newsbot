"""Deterministically import the current portfolio workbook into Route A registry YAML."""
from __future__ import annotations

import argparse
from collections import Counter
from hashlib import sha256
from pathlib import Path
import sys

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from companyk_newsbot.config import load_keyword_map  # noqa: E402
from companyk_newsbot.portfolio_registry import (  # noqa: E402
    PortfolioRegistry, display_name, normalized_identity, parse_source_name, stable_company_id,
)


def select_sheet_and_column(workbook, sheet_override: str | None, column_override: str | None) -> tuple[str, int]:
    if sheet_override:
        if sheet_override not in workbook.sheetnames: raise ValueError(f"unknown sheet: {sheet_override}")
        sheet = workbook[sheet_override]
    else:
        candidates = []
        for candidate in workbook.worksheets:
            for column in range(1, candidate.max_column + 1):
                values = [candidate.cell(row, column).value for row in range(1, candidate.max_row + 1)]
                nonblank = [str(value).strip() for value in values if value is not None and str(value).strip()]
                if 20 <= len(nonblank) <= 1000:
                    unique_ratio = len(set(nonblank)) / len(nonblank)
                    candidates.append((unique_ratio, -candidate.max_column, len(nonblank), candidate.title, column))
        if not candidates: raise ValueError("could not identify a portfolio company column")
        _, _, _, sheet_name, column = max(candidates)
        sheet = workbook[sheet_name]
        return sheet.title, column
    if column_override:
        from openpyxl.utils import column_index_from_string
        return sheet.title, column_index_from_string(column_override)
    if sheet.max_column == 1: return sheet.title, 1
    raise ValueError("column override is required for a multi-column selected sheet")


def import_registry(workbook_path: Path, *, sheet_name: str | None = None, column: str | None = None) -> tuple[PortfolioRegistry, dict[str, object]]:
    data = workbook_path.read_bytes(); digest = sha256(data).hexdigest()
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    selected_sheet, selected_column = select_sheet_and_column(workbook, sheet_name, column)
    sheet = workbook[selected_sheet]
    raw_values = [sheet.cell(row, selected_column).value for row in range(1, sheet.max_row + 1)]
    source_names = [str(value).strip() for value in raw_values if value is not None and str(value).strip()]
    duplicate_names = sorted(name for name, count in Counter(source_names).items() if count > 1)

    legacy = load_keyword_map(ROOT / "config" / "keyword_map_FINAL.yaml")
    legacy_index: dict[str, tuple[str, object]] = {}
    for legacy_name, rule in legacy.company_rules.items():
        for term in (legacy_name, *rule.aliases):
            legacy_index.setdefault(normalized_identity(term), (legacy_name, rule))

    companies, former_count, legacy_matches, ambiguous_short = [], 0, [], []
    for source_name in source_names:
        current_legal, former_legal = parse_source_name(source_name)
        current_display = display_name(current_legal)
        former_names = [former_legal] if former_legal else []
        former_count += bool(former_legal)
        search_terms = [current_display]
        if former_legal: search_terms.append(display_name(former_legal))
        match_terms = [current_display, current_legal]
        if former_legal: match_terms.extend([display_name(former_legal), former_legal])
        ambiguity = {"required_context": [], "negative_context": [], "forbidden_standalone": [],
            "required_context_for_forbidden": {}, "english_negative_context": [],
            "english_required_context_for_short_form": {}}
        matched_legacy = None
        for identity in [normalized_identity(current_display), normalized_identity(current_legal), *(normalized_identity(value) for value in former_names)]:
            if identity in legacy_index:
                matched_legacy = legacy_index[identity]; break
        if matched_legacy:
            legacy_name, rule = matched_legacy; extra = rule.model_extra or {}
            match_terms.extend(rule.aliases)
            english_short = extra.get("english_required_context_for_short_form", {})
            if not isinstance(english_short, dict): english_short = {}
            english_short = {str(key): [str(item) for item in value] for key, value in english_short.items() if isinstance(value, list)}
            forbidden_context = extra.get("required_context_for_forbidden", {})
            if not isinstance(forbidden_context, (dict, list)): forbidden_context = {}
            ambiguity = {"required_context": extra.get("required_context", []), "negative_context": extra.get("negative_terms", []),
                "forbidden_standalone": extra.get("forbidden_standalone", []),
                "required_context_for_forbidden": forbidden_context,
                "english_negative_context": extra.get("english_negative_context", []),
                "english_required_context_for_short_form": english_short}
            legacy_matches.append({"source_name": source_name, "legacy_company": legacy_name})
        if re_short_english(current_display): ambiguous_short.append(source_name)
        unique_search = unique_casefold(search_terms); unique_match = unique_casefold(match_terms)
        companies.append({"company_id": stable_company_id(current_display), "display_name": current_display,
            "source_name": source_name, "legal_names": [current_legal], "former_names": former_names,
            "search_terms": unique_search, "match_terms": unique_match, "ambiguity": ambiguity})

    modified = workbook.properties.modified
    generated_at = (modified.isoformat() if modified else "1970-01-01T00:00:00")
    registry = PortfolioRegistry.model_validate({"schema_version": "1.0", "source": {"workbook": workbook_path.name,
        "sheet": selected_sheet, "column": sheet.cell(1, selected_column).column_letter, "source_sha256": digest,
        "generated_at": generated_at, "company_count": len(companies)}, "companies": companies})
    report = {"workbook": workbook_path.name, "source_sha256": digest, "sheets": workbook.sheetnames,
        "selected_sheet": selected_sheet, "selected_column": sheet.cell(1, selected_column).column_letter,
        "raw_rows": sheet.max_row, "nonblank_company_rows": len(source_names), "company_count": len(companies),
        "former_name_count": former_count, "legacy_metadata_matches": len(legacy_matches), "legacy_matches": legacy_matches,
        "legacy_unmatched_count": len(companies) - len(legacy_matches), "duplicate_rows": duplicate_names,
        "potentially_ambiguous_short_english_names": ambiguous_short,
        "companies_with_one_search_term": sum(len(company.search_terms) == 1 for company in registry.companies)}
    return registry, report


def unique_casefold(values: list[str]) -> list[str]:
    output, seen = [], set()
    for value in values:
        value = value.strip(); key = value.casefold()
        if value and key not in seen: seen.add(key); output.append(value)
    return output


def re_short_english(value: str) -> bool:
    words = value.split()
    return value.isascii() and (len(value) <= 5 or any(len(word.strip(".,'")) <= 3 for word in words))


def write_outputs(registry: PortfolioRegistry, report: dict[str, object], registry_path: Path, report_path: Path) -> None:
    registry_path.write_text(yaml.safe_dump(registry.model_dump(mode="json"), allow_unicode=True, sort_keys=False), encoding="utf-8")
    lines = ["# Portfolio Registry Import Report", "", f"- Workbook: `{report['workbook']}`",
        f"- SHA-256: `{report['source_sha256']}`", f"- Sheets: {', '.join(report['sheets'])}",
        f"- Selected: `{report['selected_sheet']}!{report['selected_column']}1:{report['selected_column']}{report['raw_rows']}`",
        f"- Raw rows: {report['raw_rows']}", f"- Nonblank rows: {report['nonblank_company_rows']}",
        f"- Final companies: {report['company_count']}", f"- Former names: {report['former_name_count']}",
        f"- Legacy metadata matches: {report['legacy_metadata_matches']}", f"- Legacy unmatched: {report['legacy_unmatched_count']}",
        f"- Duplicate rows: {report['duplicate_rows'] or 'none'}", f"- Companies with one search term: {report['companies_with_one_search_term']}",
        "", "## Potentially ambiguous short English names", "", *(f"- {value}" for value in report['potentially_ambiguous_short_english_names']),
        "", "## Legacy metadata merges", "", *(f"- {value['source_name']} ← {value['legacy_company']}" for value in report['legacy_matches'])]
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(); parser.add_argument("workbook", type=Path); parser.add_argument("--sheet"); parser.add_argument("--column")
    parser.add_argument("--output", type=Path, default=ROOT / "config" / "portfolio_registry.yaml")
    parser.add_argument("--report", type=Path, default=ROOT / "PORTFOLIO_REGISTRY_IMPORT_REPORT.md")
    args = parser.parse_args(); registry, report = import_registry(args.workbook, sheet_name=args.sheet, column=args.column)
    write_outputs(registry, report, args.output, args.report)
    print(yaml.safe_dump(report, allow_unicode=True, sort_keys=False)); return 0


if __name__ == "__main__": raise SystemExit(main())
